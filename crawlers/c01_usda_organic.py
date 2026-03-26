"""
Crawler #1: USDA Organic Certified Operations
Source  : https://organic.ams.usda.gov/Integrity/Reports/DataHistory.aspx
Format  : "Search Results Export Format" monthly XLS snapshot
Records : ~30,000–34,000 US-only certified operations

Always stays current — every run:
  1. Selenium loads the DataHistory page (JS-rendered Angular SPA)
  2. Scrapes ALL links under "Search Results Export Format" section
  3. Sorts by year+month -> picks the MOST RECENT file automatically
  4. Downloads and parses it

When USDA publishes April, May etc. -> crawler picks it up automatically.
No hardcoded URLs. No manual updates needed.
Runtime: ~3–5 minutes
"""

import sys, os, re, csv, io, time, argparse, zipfile
from datetime import datetime

# Support running as:
#   python crawlers/c01_usda_organic.py   (from project root)
#   python c01_usda_organic.py            (from inside crawlers/)
_here = os.path.dirname(os.path.abspath(__file__))
_root = os.path.dirname(_here)
if _root not in sys.path:
    sys.path.insert(0, _root)
if _here not in sys.path:
    sys.path.insert(0, _here)

try:
    from crawlers.base_crawler import BaseCrawler
except ImportError:
    from base_crawler import BaseCrawler

DATA_HISTORY_URL = "https://organic.ams.usda.gov/Integrity/Reports/DataHistory.aspx"
SEARCH_URL       = "https://organic.ams.usda.gov/Integrity/Search"

MONTH_ORDER = {
    'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
    'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
}

US_STATES = {
    'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA','HI','ID','IL','IN',
    'IA','KS','KY','LA','ME','MD','MA','MI','MN','MS','MO','MT','NE','NV',
    'NH','NJ','NM','NY','NC','ND','OH','OK','OR','PA','RI','SC','SD','TN',
    'TX','UT','VT','VA','WA','WV','WI','WY','DC','PR','VI','GU','AS','MP',
}

NON_US_COUNTRIES = {
    'canada','mexico','germany','france','italy','spain','china','japan',
    'brazil','australia','new zealand','united kingdom','uk','gb','india',
    'argentina','chile','peru','colombia','korea','taiwan','netherlands',
    'belgium','switzerland','austria','denmark','sweden','norway','finland',
    'poland','czech republic','turkey','israel','egypt','south africa',
    'pakistan','bangladesh','thailand','vietnam','indonesia','philippines',
}

AGENT_STATE_HINTS = {
    'oregon tilth':'OR','otco':'OR','ccof':'CA','california certified':'CA',
    'midwest organic':'IA','mosa':'WI','vermont organic':'VT','nofa':'MA',
    'nofa-ny':'NY','pennsylvania certified':'PA','pcof':'PA','wsda':'WA',
    'ioia':'IA','ocia':'NE','quality assurance international':'CA','qai':'CA',
    'primus':'CA','baystate organic':'MA','ohio ecological':'OH',
}

STATUS_MAP = {
    'C':'Certified','S':'Suspended','R':'Revoked',
    'V':'Surrendered','N':'Not Certified','E':'Exempt',
}


def _month_sort_key(link_text: str, href: str) -> tuple:
    """Return (year, month) for sorting — higher = more recent."""
    text = (link_text or '').lower()
    href_l = (href or '').lower()
    combined = text + ' ' + href_l

    year_m = re.search(r'(20\d{2})', combined)
    year = int(year_m.group(1)) if year_m else datetime.utcnow().year

    for name, num in MONTH_ORDER.items():
        if name in text:
            return (year, num)

    m = re.search(r'20\d{2}[_\-]?(\d{2})', href_l)
    if m:
        return (year, int(m.group(1)))

    return (year, 0)


def _extract_agent_state(certifier: str) -> str:
    cl = (certifier or '').lower()
    for hint, st in AGENT_STATE_HINTS.items():
        if hint in cl:
            return st
    m = re.search(r'\(([A-Z]{2})\)', certifier or '')
    if m and m.group(1) in US_STATES:
        return m.group(1)
    return ''


def _parse_certified_categories(cert_prods: str) -> dict:
    cats = (cert_prods or '').lower()
    return {
        'crop_certified':      'crop'      in cats,
        'livestock_certified': 'livestock' in cats,
        'handling_certified':  'handling'  in cats,
        'wild_crop_certified': 'wild'      in cats,
    }


def _infer_operation_type(cert_prods: str) -> str:
    cats = (cert_prods or '').lower()
    parts = []
    if 'crop'      in cats: parts.append('Grower')
    if 'livestock' in cats: parts.append('Livestock Producer')
    if 'handling'  in cats: parts.append('Handler/Processor')
    if 'wild'      in cats: parts.append('Wild Crop Harvester')
    return ' + '.join(parts) if parts else 'Grower'


class UsdaOrganicCrawler(BaseCrawler):

    def __init__(self, fresh=False, csv_url=None, **kwargs):
        super().__init__(**kwargs)
        self.fresh   = fresh
        self.csv_url = csv_url

    def get_dataset_name(self):  return "usda_organic"
    def get_source_url(self):    return SEARCH_URL

    def get_niche_fields(self):
        return [
            'organic_program','operation_type','certifying_agent',
            'certifying_agent_state','certification_status','nop_id',
            'client_id','certificate_number','effective_date','expiration_date',
            'total_acreage','certified_products_raw','all_products',
            'crop_certified','livestock_certified','handling_certified','wild_crop_certified',
            'contact_name','county',
            'is_broker','is_dairy','is_retailer','is_distributor',
            'is_csa','is_poultry','is_restaurant','is_storage',
        ]

    # ── Strategy 1: Selenium scrapes the JS-rendered DataHistory page ─────

    def _selenium_get_latest_url(self) -> str:
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager
        except ImportError:
            self.logger.warning("selenium/webdriver-manager not installed — skipping")
            return ''

        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--log-level=3")
        options.add_argument("--window-size=1920,1080")

        self.logger.info("Selenium: loading DataHistory page…")
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=options
        )
        found_links = []
        try:
            driver.get(DATA_HISTORY_URL)
            # Wait until at least one anchor with an xlsx/csv/xls href appears
            try:
                WebDriverWait(driver, 25).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//a[contains(@href,'.xls') or contains(@href,'.csv') or contains(@href,'.zip')]")
                    )
                )
            except Exception:
                self.logger.warning("Timed out waiting for download links — trying anyway")

            time.sleep(2)   # let remaining JS settle

            # Grab every anchor that looks like a data download
            anchors = driver.find_elements(
                By.XPATH,
                "//a[contains(@href,'.xls') or contains(@href,'.csv') or contains(@href,'.zip')]"
            )
            self.logger.info(f"Total download anchors found: {len(anchors)}")

            for a in anchors:
                href      = a.get_attribute('href') or ''
                link_text = a.text.strip()

                # We only want "Search Results Export Format" links.
                # These link texts are month names: January, February …
                # (Data Submission Template links have the same month names but
                #  go to different files — identify by proximity to the section header)
                # Simplest heuristic: try to find the section heading in nearby DOM
                try:
                    # Walk up to the nearest table cell / div and check its text
                    parent_text = driver.execute_script("""
                        var el = arguments[0];
                        // Walk up max 6 levels
                        for (var i=0; i<6; i++) {
                            el = el.parentElement;
                            if (!el) break;
                            var t = el.innerText || '';
                            if (t.length > 20) return t;
                        }
                        return '';
                    """, a).lower()
                except Exception:
                    parent_text = ''

                is_search_results = 'search results export' in parent_text

                found_links.append({
                    'href':       href,
                    'text':       link_text,
                    'preferred':  is_search_results,
                })
                self.logger.info(
                    f"  [{link_text}] -> {href[:80]}  "
                    f"(search_results={is_search_results})"
                )

        finally:
            driver.quit()

        if not found_links:
            return ''

        # INTEGRITY_Export_YYYYMMDD.xlsx = Search Results Export (all fields one sheet) <-- WANT
        # INTEGRITY_Data_YYYYMMDD.xlsx   = Data Submission Template (split sheets)      <-- skip if possible
        # Filter to monthly files only (2026/2025 month links, not annual year files)
        monthly = [
            l for l in found_links
            if re.search(r'20\d{2}[01]\d01\.(xlsx?)', l['href'], re.I)
        ]
        pool = monthly if monthly else found_links

        # Within monthly files, strongly prefer Export over Data
        export_links = [l for l in pool if 'export' in l['href'].lower()]
        if export_links:
            pool = export_links
            self.logger.info(f"Filtered to {len(pool)} INTEGRITY_Export (Search Results) links")
        else:
            self.logger.info(f"No Export links found, using all {len(pool)} monthly links")

        # Sort by recency, pick most recent
        pool.sort(key=lambda l: _month_sort_key(l['text'], l['href']), reverse=True)
        best = pool[0]
        self.logger.info(
            f"Selected most recent link: [{best['text']}] -> {best['href']}"
        )
        return best['href']

    # ── Strategy 2: probe confirmed URL pattern (from USDA MonthlyReports) ─
    # Confirmed pattern from live site:
    #   https://organic.ams.usda.gov/Integrity/MonthlyReports/INTEGRITY_Data_YYYYMM01.xlsx
    # Try current month and 5 months back so the pipeline always finds something.

    def _probe_url_patterns(self) -> str:
        import requests
        self.logger.info("Probing confirmed USDA MonthlyReports URL patterns...")

        today   = datetime.utcnow()
        base    = "https://organic.ams.usda.gov/Integrity/MonthlyReports/"
        headers = {'User-Agent': 'Mozilla/5.0 (compatible; FirmableBot/1.0)'}

        for months_back in range(0, 7):
            m = today.month - months_back
            y = today.year
            while m <= 0:
                m += 12
                y -= 1

            # Try Export first (Search Results = all fields on one sheet), then Data fallback
            for prefix in ('INTEGRITY_Export', 'INTEGRITY_Data'):
                filename = f"{prefix}_{y}{m:02d}01.xlsx"
                url = base + filename
                try:
                    r = requests.head(url, headers=headers, timeout=10, allow_redirects=True)
                    size = int(r.headers.get('content-length', 0))
                    if r.status_code == 200 and size > 10_000:
                        self.logger.info(f"Probe hit: {url}  ({size:,} bytes)")
                        return url
                    else:
                        self.logger.debug(f"Probe miss: {url}  status={r.status_code}")
                except Exception as e:
                    self.logger.debug(f"Probe error {url}: {e}")

        return ''

    # ── crawl() ───────────────────────────────────────────────────────────

    def crawl(self) -> list:
        # Manual override (--url flag)
        url = self.csv_url

        if not url:
            # Try Selenium first (most reliable)
            url = self._selenium_get_latest_url()

        if not url:
            # Fall back to URL pattern probing
            url = self._probe_url_patterns()

        if not url:
            self.logger.error(
                "Could not auto-detect download URL.\n"
                "  Manual fix:\n"
                "  1. Open https://organic.ams.usda.gov/Integrity/Reports/DataHistory.aspx\n"
                "  2. Right-click the latest month under 'Search Results Export Format'\n"
                "  3. Copy link address\n"
                "  4. Run: python crawlers/c01_usda_organic.py --url 'PASTE_URL' --fresh"
            )
            return []

        return self._download_file(url)

    # ── Download CSV / XLS / XLSX / ZIP ───────────────────────────────────

    def _download_file(self, url: str) -> list:
        import requests

        self.logger.info(f"Downloading: {url}")
        headers = {
            'User-Agent': 'Mozilla/5.0 (compatible; FirmableBot/1.0)',
            'Referer':    DATA_HISTORY_URL,
            'Accept':     'application/octet-stream,*/*',
        }

        # Stream download in chunks — avoids timeout on large files
        for attempt in range(3):
            try:
                r = requests.get(
                    url, headers=headers,
                    timeout=(30, 600),   # (connect timeout, read timeout)
                    stream=True
                )
                r.raise_for_status()

                chunks = []
                downloaded = 0
                for chunk in r.iter_content(chunk_size=1024 * 1024):  # 1MB chunks
                    if chunk:
                        chunks.append(chunk)
                        downloaded += len(chunk)
                        if downloaded % (5 * 1024 * 1024) < 1024 * 1024:
                            self.logger.info(f"  Downloaded {downloaded / 1024 / 1024:.1f} MB...")

                content = b''.join(chunks)
                self.logger.info(f"Download complete: {len(content):,} bytes")
                break

            except Exception as e:
                self.logger.warning(f"Download attempt {attempt+1}/3 failed: {e}")
                if attempt == 2:
                    raise
                time.sleep(10)

        url_l = url.lower()

        # ── Unzip if needed ────────────────────────────────────────────
        if url_l.endswith('.zip') or (content[:2] == b'PK' and not url_l.endswith('.xlsx')):
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                names     = zf.namelist()
                preferred = [n for n in names if 'search' in n.lower() or 'export' in n.lower()]
                target    = preferred[0] if preferred else names[0]
                self.logger.info(f"Extracting '{target}' from ZIP")
                content   = zf.read(target)
                url_l     = target.lower()

        # ── Excel ──────────────────────────────────────────────────────
        if url_l.endswith('.xlsx') or url_l.endswith('.xls'):
            return self._parse_excel(content)

        # ── CSV ────────────────────────────────────────────────────────
        text   = content.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        rows   = list(reader)
        self.logger.info(f"CSV rows: {len(rows):,}  Cols: {list(reader.fieldnames or [])[:8]}")
        return rows

    def _parse_excel(self, content: bytes) -> list:
        # ── Try pandas first (10x faster on large files) ──────────────
        try:
            import pandas as pd

            self.logger.info("Parsing Excel with pandas (fast path)...")
            xl = pd.ExcelFile(io.BytesIO(content))
            sheet_names = xl.sheet_names
            self.logger.info(f"Excel sheets: {sheet_names}")

            def read_sheet(name):
                df = xl.parse(name, dtype=str)
                df = df.fillna('')
                # Strip whitespace from all string columns
                for col in df.columns:
                    df[col] = df[col].astype(str).str.strip()
                self.logger.info(f"  Sheet '{name}': {len(df):,} rows x {len(df.columns)} cols")
                self.logger.info(f"  Columns: {list(df.columns)[:15]}")
                return df

            # ── Single-sheet (Export file) ─────────────────────────────
            if len(sheet_names) == 1:
                df = read_sheet(sheet_names[0])
                return df.to_dict('records')

            # ── Multi-sheet (Data file): join Operations + Scopes ──────
            # Find the operations sheet (most rows, or named 'Operations'/'Certified')
            ops_sheet = sheet_names[0]
            scope_sheet = sheet_names[1] if len(sheet_names) > 1 else None

            # Override by name if obvious
            for s in sheet_names:
                sl = s.lower()
                if 'operation' in sl or 'certified' in sl or 'search' in sl:
                    ops_sheet = s
                    break

            self.logger.info(f"Operations sheet: '{ops_sheet}'")
            df_ops = read_sheet(ops_sheet)

            if scope_sheet and scope_sheet != ops_sheet:
                self.logger.info(f"Scopes sheet: '{scope_sheet}'")
                df_scope = read_sheet(scope_sheet)

                # Find the join key (NOP ID column) in both dataframes
                nop_cols_ops = [c for c in df_ops.columns
                                if 'nop' in c.lower() or 'operation id' in c.lower()]
                nop_cols_scope = [c for c in df_scope.columns
                                  if 'nop' in c.lower() or 'operation id' in c.lower()]

                if nop_cols_ops and nop_cols_scope:
                    key_ops   = nop_cols_ops[0]
                    key_scope = nop_cols_scope[0]
                    self.logger.info(f"Joining on: ops['{key_ops}'] = scope['{key_scope}']")

                    # Aggregate scope rows: group by NOP ID, join product strings
                    agg = {}
                    for _, row in df_scope.iterrows():
                        nid = str(row.get(key_scope, '')).strip()
                        if not nid:
                            continue
                        if nid not in agg:
                            agg[nid] = dict(row)
                        else:
                            for k, v in row.items():
                                if v and k != key_scope:
                                    existing = str(agg[nid].get(k, ''))
                                    if not existing:
                                        agg[nid][k] = v
                                    elif str(v) not in existing:
                                        agg[nid][k] = existing + ', ' + str(v)

                    # Merge into ops
                    def merge_scope(row):
                        nid = str(row.get(key_ops, '')).strip()
                        if nid in agg:
                            for k, v in agg[nid].items():
                                if k != key_scope and not row.get(k):
                                    row[k] = v
                        return row

                    df_ops = df_ops.apply(merge_scope, axis=1)
                    self.logger.info(f"Merged {len(df_ops):,} rows with scope data")
                else:
                    self.logger.warning(
                        "Could not find NOP ID column for join — "
                        "using operations sheet only"
                    )

            return df_ops.to_dict('records')

        except ImportError:
            self.logger.warning("pandas not installed — falling back to openpyxl (slower)")

        # ── openpyxl fallback ──────────────────────────────────────────
        try:
            import openpyxl
        except ImportError:
            self.logger.error("Run: pip install openpyxl  (or pip install pandas openpyxl)")
            raise

        self.logger.info("Parsing Excel with openpyxl...")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet_names = [ws.title for ws in wb.worksheets]
        self.logger.info(f"Excel sheets: {sheet_names}")

        def sheet_to_rows(ws):
            rows_iter  = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], []
            col_names = [
                str(h).strip() if h is not None else f'col_{i}'
                for i, h in enumerate(header_row)
            ]
            rows = []
            for row in rows_iter:
                rows.append({
                    col_names[i]: (str(v).strip() if v is not None else '')
                    for i, v in enumerate(row) if i < len(col_names)
                })
            return col_names, rows

        if len(wb.worksheets) == 1:
            target_ws = wb.worksheets[0]
            self.logger.info(f"Sheet '{target_ws.title}' ({target_ws.max_row:,} rows)")
            col_names, rows = sheet_to_rows(target_ws)
            self.logger.info(f"Columns: {col_names[:12]}")
            return rows

        # Multi-sheet: join ops + scopes
        self.logger.info("Multi-sheet: joining Operations + Scopes...")
        _, ops_rows = sheet_to_rows(wb.worksheets[0])
        self.logger.info(f"Ops rows: {len(ops_rows):,}")

        if len(wb.worksheets) > 1:
            _, scope_rows = sheet_to_rows(wb.worksheets[1])
            self.logger.info(f"Scope rows: {len(scope_rows):,}")
            scope_map = {}
            for sr in scope_rows:
                nid = next(
                    (sr[k] for k in sr
                     if 'nop' in k.lower() or 'operation id' in k.lower()),
                    ''
                ).strip()
                if nid:
                    if nid not in scope_map:
                        scope_map[nid] = sr
                    else:
                        for k, v in sr.items():
                            if v and not scope_map[nid].get(k):
                                scope_map[nid][k] = v
                            elif v and scope_map[nid].get(k) and str(v) not in scope_map[nid][k]:
                                scope_map[nid][k] += ', ' + str(v)

            for op in ops_rows:
                nid = next(
                    (op[k] for k in op
                     if 'nop' in k.lower() or 'operation id' in k.lower()),
                    ''
                ).strip()
                if nid and nid in scope_map:
                    op.update({k: v for k, v in scope_map[nid].items() if not op.get(k)})

        return ops_rows

    # ── parse() ───────────────────────────────────────────────────────────

    def parse(self, raw_data: list) -> list:
        if not raw_data:
            self.logger.warning("No rows to parse")
            return []

        # ── Real column names from INTEGRITY_Export_YYYYMMDD.xlsx ─────────
        # Confirmed from live file (March 2026):
        #   cert_name          = Certifying Agent name
        #   cert_url           = Certifier website
        #   cert_email         = Certifier email
        #   op_nopopid         = NOP Operation ID (unique key)
        #   op_name            = Operation / company name
        #   op_othernames      = Trade names / DBA
        #   op_clientid        = Certifier's internal client ID
        #   op_contfirstname   = Contact first name
        #   op_contlastname    = Contact last name
        #   op_status          = C / S / R / V / N
        #   op_statuseffectivedate  = Status effective date
        #   op_nopanniversarydate   = Certificate anniversary / expiry
        #   opsc_cr            = Crops scope certified (True/False)
        #   opsc_cr_ed         = Crops scope effective date
        #   cr_certifiedproducts    = Crop certified products list
        #   cr_certifiedproducts_add = Additional crop products
        #   cr_certnos         = Certificate numbers (crops)
        #   opsc_ls            = Livestock scope certified
        #   opsc_ls_ed         = Livestock scope effective date
        #   ls_certifiedproducts    = Livestock certified products
        #   opsc_wc            = Wild Crops scope certified
        #   opsc_hl            = Handling scope certified
        #   op_physStreet / op_physCity / op_physState / op_physCountry / op_physZip
        #   op_mailStreet / op_mailCity / op_mailState / op_mailCountry / op_mailZip
        #   op_phone / op_email / op_website

        def norm(k: str) -> str:
            return re.sub(r'[^a-z0-9]', '_', (k or '').lower().strip()).strip('_')

        records     = []
        seen_keys   = set()
        non_us      = 0
        sample_done = False

        for row in raw_data:
            try:
                r = {norm(k): (str(v) if v is not None else '').strip()
                     for k, v in row.items()}

                if not sample_done:
                    self.logger.info(f"All columns: {list(r.keys())}")
                    sample_done = True

                def g(*keys):
                    for k in keys:
                        v = r.get(norm(k)) or r.get(k)
                        if v and str(v).lower() not in ('','n/a','na','none','null','nan','false'):
                            return str(v).strip()
                    return ''

                # ── Skip header/description rows embedded in Excel ─────
                # The USDA Export file has 2 descriptor rows at the top:
                # Row 1: "Operation Name", "Certifier Name", ...  (column labels)
                # Row 2: "Operation's business name", ...         (descriptions)
                operation = g('op_name','op_othernames','name')
                if operation.lower() in (
                    'operation name', "operation's business name",
                    'operation name (dba)', 'name', 'business name'
                ):
                    continue

                # ── Core fields ────────────────────────────────────────
                # Confirmed column names from live file (March 2026):
                # op_name, cert_name, op_status, op_nopopid, op_clientid
                # op_statuseffectivedate, op_nopanniversarydate
                # op_physstreet, op_physcity, op_physstate, op_physcountry, op_physzip
                # op_mailstreet, op_mailcity, op_mailstate, op_mailcountry, op_mailzip
                # op_contfirstname, op_contlastname, op_phone, op_email, op_website
                # opsc_cr, opsc_ls, opsc_wc, opsc_hl (scope booleans)
                # cr_certifiedproducts, ls_certifiedproducts etc.
                # certifying_agent_state = cert_state (certifier's state)

                city      = g('oppa_city','opma_city','city')
                state     = g('oppa_state','opma_state','state')
                country   = g('oppa_country','opma_country','country')
                address   = g('oppa_line1','opma_line1','address')
                zip_code  = g('oppa_zip','opma_zip','zip')
                county    = g('oppa_county','opma_county','county')
                phone     = g('op_phone','phone')
                email     = g('op_email','email')
                website   = g('op_url','website')
                certifier = g('cert_name','certifying_agent','certifier')
                # cert_state does not exist — use operation's physical state
                cert_state= g('oppa_state','opma_state')
                status_raw= g('op_status','status')
                nop_id    = g('op_nopopid','nop_id','nopopid')
                client_id = g('op_clientid','clientid')
                eff_date  = g('op_statuseffectivedate','effective_date')
                exp_date  = g('op_nopanniversarydate','expiration_date')
                contact   = (g('op_contfirstname') + ' ' + g('op_contlastname')).strip()
                total_acreage = g('op_totalacreage','totalacreage')

                # Business type flags (extra niche fields)
                is_broker      = g('opex_broker')
                is_dairy       = g('opex_dairy')
                is_retailer    = g('opex_retailer')
                is_distributor = g('opex_distributor')
                is_csa         = g('opex_csa')
                is_poultry     = g('opex_poultry')
                is_restaurant  = g('opex_restaurant')
                is_storage     = g('opex_storage')

                # ── Scope / certified products ─────────────────────────
                crop_scope = g('opsc_cr')
                ls_scope   = g('opsc_ls')
                wc_scope   = g('opsc_wc')
                hl_scope   = g('opsc_handling')

                cr_prods   = g('cr_certifiedproducts','cr_certifiedproducts_add')
                ls_prods   = g('ls_certifiedproducts','ls_certifiedproducts_add')
                wc_prods   = g('wc_certifiedproducts','wc_certifiedproducts_add')
                hl_prods   = g('han_certifiedproducts','han_certifiedproducts_add')
                cert_nos   = g('cr_certnos','ls_certnos','wc_certnos','hand_certnos')

                # Build combined certified products string
                scope_parts = []
                if crop_scope.lower() in ('true','yes','1','t'): scope_parts.append('Crops')
                if ls_scope.lower()   in ('true','yes','1','t'): scope_parts.append('Livestock')
                if hl_scope.lower()   in ('true','yes','1','t'): scope_parts.append('Handling')
                if wc_scope.lower()   in ('true','yes','1','t'): scope_parts.append('Wild Crops')
                cert_prods   = ', '.join(scope_parts) if scope_parts else ''
                all_products = ', '.join(filter(None, [cr_prods, ls_prods, wc_prods, hl_prods]))

                if not operation:
                    continue

                # ── US-only filter ─────────────────────────────────────
                # Try all state fields in order of reliability
                state_up = (
                    state.strip().upper() or           # oppa_state (physical)
                    g('opma_state').strip().upper()     # opma_state (mailing)
                )

                # Known US certifier name keywords — if any match, it's a US op
                US_CERTIFIER_KEYWORDS = {
                    'ccof','oregon tilth','pennsylvania certified','nofa',
                    'vermont organic','baystate','mofga','midwest organic',
                    'mosa','ocia','where food comes from','global organic alliance',
                    'quality certification services','qai','quality assurance int',
                    'nature\'s international','onecert','ohio ecological',
                    'idaho state','washington state dept','montana dept',
                    'texas dept','colorado dept','kentucky dept','iowa dept',
                    'maryland dept','oregon dept','new mexico dept',
                    'indiana state','georgia crop','clemson','minnesota crop',
                    'oklahoma dept','wisconsin dept','wyoming dept',
                    'utah dept','virginia dept','south carolina',
                    'natural food certif','onmark','pro-cert',
                    'otco','tilth','paorganic','wfcforganic',
                }
                cert_lo = certifier.lower()
                is_us_certifier = any(kw in cert_lo for kw in US_CERTIFIER_KEYWORDS)

                if state_up and state_up in US_STATES:
                    pass  # physical/mailing state is a US state -> keep
                elif is_us_certifier:
                    pass  # known US certifier -> keep
                else:
                    non_us += 1
                    continue

                # ── Niche processing ───────────────────────────────────
                # Use best available state for storage
                if not state_up:
                    state_up = g('opma_state').strip().upper()
                status_clean = STATUS_MAP.get(status_raw.upper()[:1], status_raw) or 'Certified'
                op_type      = _infer_operation_type(cert_prods)
                cats         = _parse_certified_categories(cert_prods)
                agent_state  = _extract_agent_state(certifier)

                dedup_key = (operation.lower(), state_up, nop_id or client_id)
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)

                records.append({
                    'company_name':           operation,
                    'city':                   city,
                    'state':                  state_up,
                    'country':                'United States',
                    'county':                 county,
                    'address':                address,
                    'zip_code':               zip_code,
                    'phone':                  phone,
                    'email':                  email,
                    'website':                website,
                    'contact_name':           contact,
                    # Niche fields
                    'organic_program':        'NOP',
                    'operation_type':         op_type,
                    'certifying_agent':       certifier,
                    'certifying_agent_state': agent_state,
                    'certification_status':   status_clean,
                    'certificate_number':     cert_nos,
                    'client_id':              client_id,
                    'effective_date':         eff_date,
                    'expiration_date':        exp_date,
                    'total_acreage':          total_acreage,
                    'certified_products_raw': cert_prods,
                    'all_products':           all_products[:500] if all_products else '',
                    'crop_certified':         str(cats['crop_certified']),
                    'livestock_certified':    str(cats['livestock_certified']),
                    'handling_certified':     str(cats['handling_certified']),
                    'wild_crop_certified':    str(cats['wild_crop_certified']),
                    'nop_id':                 nop_id,
                    # Business type flags
                    'is_broker':              is_broker,
                    'is_dairy':               is_dairy,
                    'is_retailer':            is_retailer,
                    'is_distributor':         is_distributor,
                    'is_csa':                 is_csa,
                    'is_poultry':             is_poultry,
                    'is_restaurant':          is_restaurant,
                    'is_storage':             is_storage,
                })

            except Exception as ex:
                self.stats['total_errors'] += 1
                self.logger.debug(f"Row parse error: {ex}")

        self.logger.info(
            f"Parsed {len(records):,} US records  ({non_us} non-US filtered out)"
        )
        return records

    # ── _store override ───────────────────────────────────────────────────

    def _store(self, records):
        if self.fresh:
            self._drop_table()
        return super()._store(records)

    def _drop_table(self):
        import sqlite3
        if not os.path.exists(self.db_path):
            return
        conn = sqlite3.connect(self.db_path)
        conn.execute(f'DROP TABLE IF EXISTS "{self.get_dataset_name()}"')
        conn.commit()
        conn.close()
        self.logger.info("Old table dropped (--fresh mode)")


# ── Entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="USDA Organic — always fetches the latest monthly snapshot automatically"
    )
    parser.add_argument("--fresh", action="store_true",
                        help="Drop existing table before inserting")
    parser.add_argument("--url", default=None, metavar="URL",
                        help="Manually supply download URL (skips auto-detect)")
    args = parser.parse_args()

    os.makedirs("logs", exist_ok=True)
    crawler = UsdaOrganicCrawler(fresh=args.fresh, csv_url=args.url)
    stats   = crawler.run()

    print(f"\n{'='*45}")
    print(f"  New records     : {stats['total_new']}")
    print(f"  Updated records : {stats['total_updated']}")
    print(f"  Duplicates skip : {stats['total_duplicates']}")
    print(f"  Errors          : {stats['total_errors']}")
    print(f"{'='*45}")

    if stats['total_new'] + stats['total_updated'] > 0:
        path = crawler.export_csv()
        if path:
            print(f"  CSV -> {path}")
    else:
        print("  ⚠  0 records saved")
        print("  Manual fix:")
        print("  1. Open: https://organic.ams.usda.gov/Integrity/Reports/DataHistory.aspx")
        print("  2. Right-click latest month under 'Search Results Export Format'")
        print("  3. Copy link address")
        print("  4. Run: python crawlers/c01_usda_organic.py --url 'PASTE_URL' --fresh")